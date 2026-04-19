import hashlib
import math
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from PIL import Image

# ====== 使用说明 ======
# 1. 把 INPUT_XLSX 改成你的题库路径
# 2. 直接运行本脚本
# 3. 会生成一个新的 xlsx 文件，不覆盖原文件

INPUT_XLSX = r"D:\学习资料\大四\毕业论文\MVTec_AD_Thesis\05_metadata\MVTec_实验题库_完整版_修订版.xlsx"
OUTPUT_XLSX = r"D:\学习资料\大四\毕业论文\MVTec_AD_Thesis\05_metadata\MVTec_实验题库_完整版_解释优化版.xlsx"
SHEET_NAME = "题库总表"


def stable_hash_int(text: str) -> int:
    return int(hashlib.md5(text.encode("utf-8")).hexdigest(), 16)


def safe_open_image(path_str: str):
    if not path_str:
        return None
    path = Path(path_str)
    if not path.exists():
        return None
    return Image.open(path)


def load_mask(mask_path: str):
    img = safe_open_image(mask_path)
    if img is None:
        return None
    arr = np.array(img)
    if arr.ndim == 3:
        arr = arr[:, :, 0]
    return (arr > 0).astype(np.uint8)


def compute_mask_stats(mask_path: str, image_path: str):
    img = safe_open_image(image_path)
    if img is None:
        return {
            "width": None,
            "height": None,
            "pixels": 0,
            "ratio": 0.0,
            "bbox": None,
            "location": "中心区域",
        }

    width, height = img.size
    mask = load_mask(mask_path)
    if mask is None:
        return {
            "width": width,
            "height": height,
            "pixels": 0,
            "ratio": 0.0,
            "bbox": None,
            "location": "中心区域",
        }

    ys, xs = np.where(mask > 0)
    if len(xs) == 0:
        return {
            "width": width,
            "height": height,
            "pixels": 0,
            "ratio": 0.0,
            "bbox": None,
            "location": "中心区域",
        }

    x1, x2 = int(xs.min()), int(xs.max())
    y1, y2 = int(ys.min()), int(ys.max())
    pixels = int(mask.sum())
    ratio = pixels / float(width * height) * 100
    cx = (x1 + x2) / 2 / width
    cy = (y1 + y2) / 2 / height

    if cy < 0.33:
        v = "上部"
    elif cy > 0.66:
        v = "下部"
    else:
        v = "中部"

    if cx < 0.33:
        h = "左侧"
    elif cx > 0.66:
        h = "右侧"
    else:
        h = "中央"

    if h == "中央" and v == "中部":
        location = "中心区域"
    elif h == "中央":
        location = v
    elif v == "中部":
        location = h
    else:
        location = f"{v}{h}"

    return {
        "width": width,
        "height": height,
        "pixels": pixels,
        "ratio": ratio,
        "bbox": (x1, y1, x2, y2),
        "location": location,
    }


def defect_desc(defect_type: str):
    mapping = {
        "污染": "污染/附着异常",
        "大破损": "明显破损",
        "小破损": "轻微破损",
        "划痕": "划痕",
        "裂缝": "裂缝",
        "针孔": "针孔",
        "色差": "局部色差",
        "挤压变形": "挤压变形",
        "弯曲": "弯曲变形",
        "翻转": "朝向异常",
        "无缺陷": "未见明显异常",
    }
    return mapping.get(str(defect_type).strip(), str(defect_type).strip() or "异常")


def subject_desc(category: str):
    return {
        "瓶子": "瓶体",
        "胶囊": "胶囊表面",
        "金属螺母": "螺母表面与边缘",
    }.get(category, "样本表面")


def confidence_value(item_id: str, ai_suggestion: str, ai_correct: str, ratio: float):
    base = stable_hash_int(item_id) % 7
    if ai_suggestion.startswith("NG"):
        if str(ai_correct).startswith("正确"):
            conf = 86 + min(10, int(ratio * 1.8)) + base % 3
        else:
            conf = 61 + base
    else:
        if str(ai_correct).startswith("正确"):
            conf = 85 + base
        else:
            conf = 58 + min(8, int(ratio * 1.5)) + base % 3
    return max(51, min(98, conf))


def build_metric_text(row, stats):
    ai_suggestion = str(row["AI建议"])
    ai_correct = str(row["AI是否正确"])
    item_id = str(row["图片ID"])
    ratio = stats["ratio"]
    pixels = stats["pixels"]
    loc = stats["location"]
    conf = confidence_value(item_id, ai_suggestion, ai_correct, ratio)

    if ai_suggestion.startswith("NG"):
        salience = "高" if ratio >= 3 else "中" if ratio >= 1 else "低"
        return f"置信度：{conf}%｜异常面积占比：{ratio:.2f}%｜异常像素：{pixels:,} px｜主异常位置：{loc}｜异常显著性：{salience}"
    else:
        consistency = "高" if str(ai_correct).startswith("正确") else "中"
        return f"置信度：{conf}%｜异常面积占比：{ratio:.2f}%｜异常像素：{pixels:,} px｜整体一致性：{consistency}｜判定位置：全局扫描"


def build_reason_text(row, stats):
    category = str(row["产品类别"])
    defect_type = str(row["缺陷类型"])
    true_label = str(row["真实标签"])
    ai_suggestion = str(row["AI建议"])
    ai_correct = str(row["AI是否正确"])
    ratio = stats["ratio"]
    loc = stats["location"]
    subj = subject_desc(category)
    defect = defect_desc(defect_type)

    if ai_suggestion.startswith("NG"):
        if true_label.startswith("NG"):
            return f"系统在{loc}检测到{defect}，{subj}与正常模板相比存在可分辨差异，异常区域约占图像{ratio:.2f}%。综合异常位置、面积与形态特征，判定为不合格。"
        return f"系统在局部扫描中发现与正常模板不一致的可疑区域，主要表现为纹理或边缘响应异常，因此输出不合格。该题为误报样本，解释反映的是AI当时的判断依据。"

    if true_label.startswith("OK"):
        return f"系统未检测到稳定且连续的异常区域，{subj}整体轮廓、纹理或表面状态处于可接受波动范围内，因此判定为合格。"

    return f"系统未检测到足够显著的异常信号，当前可见偏差规模较小（参考异常面积约{ratio:.2f}%）且不具备稳定连续性，因此仍输出合格。该题为漏检样本，解释反映的是AI当时的判断依据。"


def build_exp1_text(row, stats):
    ai_suggestion = str(row["AI建议"])
    category = str(row["产品类别"])
    subj = subject_desc(category)
    if ai_suggestion.startswith("NG"):
        return f"系统检测到{stats['location']}存在可疑异常特征，{subj}与正常样本存在差异，综合判断为不合格。"
    return f"系统未检测到稳定异常区域，{subj}整体特征处于可接受范围，综合判断为合格。"


def main():
    wb = load_workbook(INPUT_XLSX)
    ws = wb[SHEET_NAME]
    header = [c.value for c in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(header)}

    required = [
        "图片ID", "产品类别", "缺陷类型", "真实标签", "AI建议", "AI是否正确",
        "ground_truth文件路径", "图片源路径", "实验一-统一解释内容",
        "实验二-指标型解释内容", "实验二-理由型解释内容"
    ]
    missing = [x for x in required if x not in col]
    if missing:
        raise ValueError(f"缺少必要列: {missing}")

    for r in range(2, ws.max_row + 1):
        row = {name: ws.cell(r, c).value for name, c in col.items()}
        stats = compute_mask_stats(str(row.get("ground_truth文件路径") or ""), str(row.get("图片源路径") or ""))
        ws.cell(r, col["实验一-统一解释内容"]).value = build_exp1_text(row, stats)
        ws.cell(r, col["实验二-指标型解释内容"]).value = build_metric_text(row, stats)
        ws.cell(r, col["实验二-理由型解释内容"]).value = build_reason_text(row, stats)

    wb.save(OUTPUT_XLSX)
    print(f"已生成：{OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
