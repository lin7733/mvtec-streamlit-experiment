import os
import shutil
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_DIR = Path("05_metadata")
TARGET_BASENAME = "AI_质检实验题库_实验一实验二_多目标采购版.xlsx"
WORKBOOK_NAME = (
    TARGET_BASENAME
    if (WORKBOOK_DIR / TARGET_BASENAME).exists()
    else next(
        name
        for name in os.listdir(WORKBOOK_DIR)
        if name.startswith("AI_") and name.endswith(".xlsx") and ".bak" not in name.lower()
    )
)
WORKBOOK_PATH = WORKBOOK_DIR / WORKBOOK_NAME
BACKUP_PATH = WORKBOOK_DIR / f"{WORKBOOK_PATH.stem}.bak.xlsx"

QUALITY_GATE = 70
QUALITY_WEIGHT = 0.70
COST_WEIGHT = 0.30
PURCHASE_THRESHOLD = 74

PRICE_LADDERS = {
    "bottle": [6.8, 7.5, 8.3, 9.2, 10.4, 11.7, 12.9, 14.2],
    "capsule": [1.6, 1.9, 2.2, 2.6, 3.0, 3.4, 3.8, 4.3],
    "metal_nut": [0.8, 0.95, 1.1, 1.3, 1.5, 1.7, 1.9, 2.2],
}

COST_SCORES = [95, 86, 78, 70, 61, 50, 38, 25]

PRODUCT_CN = {
    "bottle": "瓶子",
    "capsule": "胶囊",
    "metal_nut": "金属螺母",
}

SEVERE_DEFECTS = {"大破损", "裂缝", "挤压变形", "弯曲"}


def stable_hash_int(text: str) -> int:
    import hashlib

    return int(hashlib.md5(text.encode("utf-8")).hexdigest(), 16)


def clamp(value: int, lo: int, hi: int) -> int:
    return max(lo, min(hi, value))


def pick_from_range(seed: str, lo: int, hi: int) -> int:
    return lo + stable_hash_int(seed) % (hi - lo + 1)


def build_headers(ws):
    headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    return headers, {header: i + 1 for i, header in enumerate(headers) if header}


def source_rows(ws):
    headers, idx = build_headers(ws)
    rows = []
    for r in range(4, ws.max_row + 1):
        row = {header: ws.cell(r, idx[header]).value for header in idx}
        row["_sheet_row"] = r
        rows.append(row)
    return headers, idx, rows


def assign_price_ranks(rows):
    by_category = defaultdict(list)
    for row in rows:
        by_category[str(row["产品类别"])].append(row)

    rank_by_item = {}
    for category, category_rows in by_category.items():
        rank_pool = [0, 1, 2, 3, 4, 5, 6, 7] * 2
        rank_pool.sort(key=lambda rank: stable_hash_int(f"{category}|price_rank|{rank}"))
        category_rows_sorted = sorted(
            category_rows,
            key=lambda row: stable_hash_int(f"{row['图片ID']}|{row['题号']}|price_slot"),
        )
        for row, rank in zip(category_rows_sorted, rank_pool):
            rank_by_item[row["图片ID"]] = rank
    return rank_by_item


def quality_score_for_row(row, price_rank, ok_conflict_ids, ng_conflict_ids):
    item_id = str(row["图片ID"])
    label = str(row["真实标签"])
    complexity = str(row["复杂度"])
    defect = str(row["缺陷类型"])

    if label.startswith("OK"):
        if item_id in ok_conflict_ids:
            return pick_from_range(item_id + "|ok_conflict_quality", 78, 88)
        if complexity == "低复杂":
            return pick_from_range(item_id + "|ok_low_quality", 82, 95)
        return pick_from_range(item_id + "|ok_high_quality", 70, 88)

    severe = defect in SEVERE_DEFECTS
    if item_id in ng_conflict_ids:
        return pick_from_range(item_id + "|ng_conflict_quality", 56, 65)
    if complexity == "低复杂" and severe:
        return pick_from_range(item_id + "|ng_low_severe", 20, 38)
    if complexity == "低复杂":
        return pick_from_range(item_id + "|ng_low_moderate", 30, 48)
    if severe:
        return pick_from_range(item_id + "|ng_high_severe", 35, 52)
    return pick_from_range(item_id + "|ng_high_moderate", 48, 65)


def price_and_cost(category, price_rank):
    price = PRICE_LADDERS[category][price_rank]
    cost_score = COST_SCORES[price_rank]
    return price, cost_score


def purchase_truth(quality_score, cost_score):
    gate_pass_code = int(quality_score >= QUALITY_GATE)
    weighted_score = round(quality_score * QUALITY_WEIGHT + cost_score * COST_WEIGHT, 1)
    purchase_gt_code = int(gate_pass_code == 1 and weighted_score >= PURCHASE_THRESHOLD)
    return gate_pass_code, weighted_score, purchase_gt_code


def choose_conflict_samples(rows, rank_by_item):
    ok_conflict_ids = set()
    ng_conflict_ids = set()
    by_category_label = defaultdict(list)
    for row in rows:
        key = (str(row["产品类别"]), str(row["真实标签"]))
        by_category_label[key].append(row)

    for category in PRICE_LADDERS:
        ok_rows = by_category_label[(category, "OK（合格）")]
        ng_rows = by_category_label[(category, "NG（不合格）")]

        ok_rows = sorted(
            ok_rows,
            key=lambda row: (
                -rank_by_item[row["图片ID"]],
                stable_hash_int(str(row["图片ID"]) + "|ok_conflict"),
            ),
        )
        ng_rows = sorted(
            ng_rows,
            key=lambda row: (
                rank_by_item[row["图片ID"]],
                stable_hash_int(str(row["图片ID"]) + "|ng_conflict"),
            ),
        )

        ok_conflict_ids.update(str(row["图片ID"]) for row in ok_rows[:2])
        ng_conflict_ids.update(str(row["图片ID"]) for row in ng_rows[:2])

    return ok_conflict_ids, ng_conflict_ids


def choose_ai_wrong_ids(rows):
    by_group = defaultdict(list)
    for row in rows:
        by_group[(row["产品类别"], row["source_true_label"])].append(row)

    ai_wrong_ids = set()
    for group_rows in by_group.values():
        low_rows = [row for row in group_rows if row["复杂度"] == "低复杂"]
        high_rows = [row for row in group_rows if row["复杂度"] == "高复杂"]
        low_rows.sort(key=lambda row: stable_hash_int(str(row["图片ID"]) + "|ai_wrong_low"))
        high_rows.sort(key=lambda row: stable_hash_int(str(row["图片ID"]) + "|ai_wrong_high"))
        ai_wrong_ids.add(str(low_rows[0]["图片ID"]))
        ai_wrong_ids.add(str(high_rows[0]["图片ID"]))
    return ai_wrong_ids


def decision_zone(label, purchase_gt_code, gate_pass_code, weighted_score, cost_score):
    if label.startswith("OK") and purchase_gt_code == 0:
        return "conflict_high_quality_high_cost"
    if label.startswith("NG") and cost_score >= 85:
        return "conflict_low_quality_low_cost"
    if purchase_gt_code == 1 and abs(weighted_score - PURCHASE_THRESHOLD) <= 3:
        return "borderline_buy"
    if purchase_gt_code == 0 and gate_pass_code == 1 and abs(weighted_score - PURCHASE_THRESHOLD) <= 3:
        return "borderline_not_buy"
    if purchase_gt_code == 1:
        return "clear_buy"
    return "clear_not_buy"


def suggestion_type(purchase_gt_code, ai_suggestion_code):
    if ai_suggestion_code == purchase_gt_code:
        return "correct_buy" if purchase_gt_code == 1 else "correct_not_buy"
    return "false_reject" if purchase_gt_code == 1 else "false_accept"


def price_level_desc(cost_score):
    if cost_score >= 85:
        return "成本优势明显"
    if cost_score >= 65:
        return "成本处于中等水平"
    return "成本压力较高"


def metric_text(
    product_cn,
    quality_score,
    supplier_price,
    cost_score,
    weighted_score,
    purchase_label,
    decision_zone,
    ai_correct_code,
):
    zone_note = {
        "conflict_high_quality_high_cost": "属于“高品质但高成本”冲突样本",
        "conflict_low_quality_low_cost": "属于“低品质但低成本”冲突样本",
        "borderline_buy": "属于接近阈值的边界采购样本",
        "borderline_not_buy": "属于接近阈值的边界放弃样本",
    }.get(decision_zone, "属于常规决策样本")
    if ai_correct_code == 0:
        return (
            f"{product_cn}质量分={quality_score}；供应价格={supplier_price:.2f}元/件；"
            f"成本分={cost_score}；AI 当前建议={purchase_label}。"
            f"该题为 AI 易错样本，系统更侧重局部风险或价格信号。{zone_note}。"
        )
    return (
        f"{product_cn}质量分={quality_score}；供应价格={supplier_price:.2f}元/件；"
        f"成本分={cost_score}；质量门槛={QUALITY_GATE}；综合分={weighted_score:.1f}；"
        f"采购阈值={PURCHASE_THRESHOLD}。建议：{purchase_label}。{zone_note}。"
    )


def reason_text(
    product_cn,
    defect_type,
    quality_score,
    supplier_price,
    cost_score,
    weighted_score,
    purchase_label,
    decision_zone,
    ai_correct_code,
):
    if ai_correct_code == 0 and purchase_label == "采购":
        return (
            f"AI 认为该{product_cn}虽然仍可见“{defect_type}”相关异常线索，但当前报价为 "
            f"{supplier_price:.2f} 元/件，价格优势较明显，因此更倾向给出采购建议。"
            f"这是一道多目标权衡下的易错样本。"
        )
    if ai_correct_code == 0 and purchase_label == "不采购":
        return (
            f"AI 对该{product_cn}采取了更保守的判断：虽然当前质量分为 {quality_score}，"
            f"但系统认为其局部风险或成本压力仍需优先关注，因此给出不采购建议。"
            f"这是一道多目标权衡下的易错样本。"
        )
    if decision_zone == "conflict_high_quality_high_cost":
        return (
            f"该{product_cn}当前质量分为 {quality_score}，已达到质量门槛，但供应价格为 "
            f"{supplier_price:.2f} 元/件，成本分仅为 {cost_score}。在质量达标的前提下，"
            f"较高报价拉低了综合分至 {weighted_score:.1f}，低于采购阈值，因此不建议采购。"
        )
    if decision_zone == "conflict_low_quality_low_cost":
        return (
            f"该{product_cn}报价仅 {supplier_price:.2f} 元/件，成本上很有吸引力，但当前质量分仅为 "
            f"{quality_score}，未达到 {QUALITY_GATE} 分质量门槛，仍表现出“{defect_type}”相关风险。"
            f"即使价格较低，也不建议采购。"
        )
    if purchase_label == "采购":
        return (
            f"该{product_cn}质量分为 {quality_score}，达到采购所需的质量门槛；结合 "
            f"{supplier_price:.2f} 元/件的报价，其{price_level_desc(cost_score)}，综合分为 "
            f"{weighted_score:.1f}，高于采购阈值，因此建议采购。"
        )
    if quality_score < QUALITY_GATE:
        return (
            f"该{product_cn}当前质量分仅为 {quality_score}，未达到 {QUALITY_GATE} 分门槛，说明样本仍存在"
            f"“{defect_type}”对应的质量隐患。即使报价为 {supplier_price:.2f} 元/件，也不建议采购。"
        )
    return (
        f"该{product_cn}虽然质量分达到 {quality_score}，但报价为 {supplier_price:.2f} 元/件，"
        f"{price_level_desc(cost_score)}，综合分仅为 {weighted_score:.1f}，低于采购阈值，因此不建议采购。"
    )


def note_text(decision_zone, ai_correct_code):
    note = {
        "conflict_high_quality_high_cost": "高品质但高成本冲突样本",
        "conflict_low_quality_low_cost": "低品质但低成本冲突样本",
        "borderline_buy": "边界采购样本",
        "borderline_not_buy": "边界放弃样本",
        "clear_buy": "常规采购样本",
        "clear_not_buy": "常规不采购样本",
    }[decision_zone]
    if ai_correct_code == 1 and decision_zone.startswith(("conflict", "borderline")):
        note += "；适合研究适度依赖"
    return note


def main():
    shutil.copy2(WORKBOOK_PATH, BACKUP_PATH)

    wb = load_workbook(WORKBOOK_PATH)
    exp1_ws = wb.worksheets[2]
    exp2_ws = wb.worksheets[3]

    _, _, exp1_rows = source_rows(exp1_ws)
    exp2_headers, exp2_idx = build_headers(exp2_ws)

    rank_by_item = assign_price_ranks(exp1_rows)
    ok_conflict_ids, ng_conflict_ids = choose_conflict_samples(exp1_rows, rank_by_item)

    generated_rows = []
    for row in exp1_rows:
        category = str(row["产品类别"])
        product_cn = str(row["产品类别中文"]) or PRODUCT_CN.get(category, category)
        label = str(row["真实标签"])
        item_id = str(row["图片ID"])
        price_rank = rank_by_item[item_id]
        quality_score = quality_score_for_row(row, price_rank, ok_conflict_ids, ng_conflict_ids)
        supplier_price, cost_score = price_and_cost(category, price_rank)
        gate_pass_code, weighted_score, purchase_gt_code = purchase_truth(quality_score, cost_score)
        purchase_gt_label = "采购" if purchase_gt_code == 1 else "不采购"
        zone = decision_zone(label, purchase_gt_code, gate_pass_code, weighted_score, cost_score)

        generated_rows.append(
            {
                "题号": row["题号"],
                "图片ID": item_id,
                "产品类别": category,
                "产品类别中文": product_cn,
                "缺陷类型": row["缺陷类型"],
                "复杂度": row["复杂度"],
                "复杂度代码": row["复杂度代码"],
                "图片源路径": row["图片源路径"],
                "source_true_label": label,
                "quality_score": quality_score,
                "supplier_price": supplier_price,
                "cost_score": cost_score,
                "quality_gate": QUALITY_GATE,
                "quality_weight": QUALITY_WEIGHT,
                "cost_weight": COST_WEIGHT,
                "purchase_threshold": PURCHASE_THRESHOLD,
                "gate_pass_code": gate_pass_code,
                "weighted_score": weighted_score,
                "purchase_gt_code": purchase_gt_code,
                "purchase_gt_label": purchase_gt_label,
                "decision_zone": zone,
            }
        )

    ai_wrong_ids = choose_ai_wrong_ids(generated_rows)

    for row in generated_rows:
        purchase_gt_code = row["purchase_gt_code"]
        ai_suggestion_code = 1 - purchase_gt_code if row["图片ID"] in ai_wrong_ids else purchase_gt_code
        ai_correct_code = int(ai_suggestion_code == purchase_gt_code)
        ai_label = "采购" if ai_suggestion_code == 1 else "不采购"
        row["ai_suggestion_code"] = ai_suggestion_code
        row["ai_suggestion_label"] = ai_label
        row["suggestion_type"] = suggestion_type(purchase_gt_code, ai_suggestion_code)
        row["ai_correct_code"] = ai_correct_code
        row["ai_correct_label"] = "正确" if ai_correct_code == 1 else "错误"
        row["实验二-指标型解释内容"] = metric_text(
            row["产品类别中文"],
            row["quality_score"],
            row["supplier_price"],
            row["cost_score"],
            row["weighted_score"],
            ai_label,
            row["decision_zone"],
            ai_correct_code,
        )
        row["实验二-理由型解释内容"] = reason_text(
            row["产品类别中文"],
            row["缺陷类型"],
            row["quality_score"],
            row["supplier_price"],
            row["cost_score"],
            row["weighted_score"],
            ai_label,
            row["decision_zone"],
            ai_correct_code,
        )
        row["备注"] = note_text(row["decision_zone"], ai_correct_code)

    for sheet_row, row in enumerate(generated_rows, start=4):
        for header in exp2_headers:
            if header is None:
                continue
            exp2_ws.cell(sheet_row, exp2_idx[header]).value = row.get(header)

    wb.save(WORKBOOK_PATH)

    purchase_counter = defaultdict(int)
    zone_counter = defaultdict(int)
    ai_correct_total = 0
    for row in generated_rows:
        purchase_counter[row["purchase_gt_label"]] += 1
        zone_counter[row["decision_zone"]] += 1
        ai_correct_total += row["ai_correct_code"]

    print(f"Updated workbook: {WORKBOOK_PATH}")
    print(f"Backup created: {BACKUP_PATH}")
    print("Purchase label distribution:", dict(sorted(purchase_counter.items())))
    print("Decision zone distribution:", dict(sorted(zone_counter.items())))
    print(f"AI correct: {ai_correct_total}/{len(generated_rows)} = {ai_correct_total / len(generated_rows):.2%}")


if __name__ == "__main__":
    main()
